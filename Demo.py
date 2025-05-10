
import os
import time
import json
import pandas as pd
from youtube_transcript_api import YouTubeTranscriptApi, TranscriptsDisabled, NoTranscriptFound
from pytube import YouTube
from moviepy import VideoFileClip
import google.generativeai as genai
from transformers import pipeline
from openpyxl import load_workbook
import csv
from io import StringIO
from google.generativeai.types import GenerationConfig


# ====== CONFIG =======
GEMINI_API_KEY = "AIzaSyDuxCYXv_OZUcScYlOuwS4M6XaJJEg4OGk"
EXCEL_INPUT_FILE = r"C:\Users\parth\OneDrive\Documents\IDS 506 Mental health\.venv\YoutubeLinks.xlsx"
EXCEL_INPUT_SHEET = "Sheet1"
EXCEL_OUTPUT_FILE = r"C:\Users\parth\OneDrive\Documents\IDS 506 Mental health\.venv\New_Book.xlsx"
EXCEL_OUTPUT_SHEET = "Sheet1"
# =====================

genai.configure(api_key=GEMINI_API_KEY)
emotion_classifier = pipeline("text-classification", model="nateraw/bert-base-uncased-emotion", return_all_scores=True)

# Global constant: EXACTLY 124 columns in this order
COLUMNS_124 = ["YouTube URL", "Gender", "Age Group", "Race", "Mental Health Illness: Anxiety disorders", "Mental Health Illness: Psychotic disorders", "Mental Health Illness: Eating disorders", 
"Mental Health Illness: Impulse control and addiction disorders", "Mental Health Illness: Personality disorders", "Mental Health Illness: Obsessive-compulsive disorder (OCD)", "Mental Health Illness: Post-traumatic stress disorder (PTSD)", "Mental Health Illness: Schizophrenia",


"Emotion Score: Joy", "Emotion Score: Sadness", "Emotion Score: Anger", "Emotion Score: Fear", "Emotion Score: Surprise",

"Support: Family", "Support: Friends", "Support: Online Communities", "Support: Community Support Services", "Support: Religious or Spiritual Community", "Support: Pets/Emotional Support Animals",

"Stigma: Self", "Stigma: Social", "Stigma: Family", "Stigma: Institutional", "Stigma: Religious/Cultural",

"Coping: Lifestyle Changes", "Coping: Mindfulness and Meditation", "Coping: Hobbies", "Coping: Problem-Solving", "Coping: Emotional Suppression", "Coping: Aggression", "Coping: Social Withdrawal",

"Treatment: Psychotherapy", "Treatment: Medication", "Treatment: Hospitalization", "Treatment: Brain Stimulation Therapies", "Treatment: Complementary & Alternative Medicine", "Treatment: Trauma-Focused Therapies",

"Cause: Trauma and Abuse", "Cause: Grief and Loss", "Cause: Social and Interpersonal Factors", "Cause: Academic and Occupational Stress", "Cause: Financial Strain", "Cause: Legal Issues", "Cause: Biological and Health Factors", "Cause: Substance Use",

"Recovery Status",

"PHQ9: Little Interest", "PHQ9: Feeling Down", "PHQ9: Sleep Problems", "PHQ9: Low Energy", "PHQ9: Appetite Issues", "PHQ9: Low Self-Worth", "PHQ9: Trouble Concentrating", "PHQ9: Restlessness", "PHQ9: Suicidal Thoughts",

"GAD7: Feeling Nervous", "GAD7: Can’t Control Worrying", "GAD7: Worrying Too Much", "GAD7: Trouble Relaxing", "GAD7: Restlessness", "GAD7: Easily Annoyed", "GAD7: Fear Something Awful",
"Hook Present"]

def get_emotion_scores(transcript):
    results = emotion_classifier(transcript[:512])
    top_emotions = sorted(results[0], key=lambda x: x['score'], reverse=True)
    scores = {e['label']: round(e['score'], 3) for e in top_emotions}
    return top_emotions[0]['label'], scores

def load_video_urls_from_excel(file_path: str, sheet_name: str = "Sheet1") -> list:
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    return df.iloc[:, 0].dropna().tolist()

def append_to_excel(output_file: str, data_row: list):
    """Append a single analysis row to the master Excel sheet, padding / truncating
    so it lines up with our  ⟨columns⟩ definition below."""

    columns = COLUMNS_124




    # --- pad / trim row to match columns ---
    if len(data_row) > len(columns):
        data_row = data_row[: len(columns)]
    elif len(data_row) < len(columns):
        data_row += ["Not mentioned"] * (len(columns) - len(data_row))

    df_new = pd.DataFrame([data_row], columns=columns)

    # --- write (create if needed) ---
    if not os.path.exists(output_file):
        df_new.to_excel(output_file, index=False, sheet_name=EXCEL_OUTPUT_SHEET)
        print(f"✅ Created {output_file} and wrote first row → {EXCEL_OUTPUT_SHEET}")
        return

    book = load_workbook(output_file)
    if EXCEL_OUTPUT_SHEET not in book.sheetnames:
        start_row = 0
        header = True
    else:
        start_row = book[EXCEL_OUTPUT_SHEET].max_row
        header = False

    with pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df_new.to_excel(
            writer,
            index=False,
            sheet_name=EXCEL_OUTPUT_SHEET,
            header=header,
            startrow=start_row,
        )
    print(f"✅ Appended analysis row to {EXCEL_OUTPUT_SHEET} in {output_file}")


def get_youtube_transcript(video_id):
    transcript = YouTubeTranscriptApi.get_transcript(video_id)
    return " ".join([entry["text"] for entry in transcript])

def analyze_transcript(transcript,url):
    
    prompt = f"""

You are a professional Mental Health Specialist. Your task is to evaluate a person's Post Mental Health Recovery experience based on the provided YouTube transcript.

Please return your response strictly as a *JSON object* with exactly *61 key-value pairs*. Each key must exactly match one of the predefined field names below. Each value must be a string.

If the transcript does not mention or imply a specific detail, the corresponding value must be "Not mentioned" (case-sensitive). Do not omit any keys. Do not return any extra text or explanation — only return a clean JSON object.

Here is the exact list of keys (61 total), and your output must include them in this order:

[ "Gender", "Age Group", "Race", "Mental Health Illness: Anxiety disorders", "Mental Health Illness: Psychotic disorders", "Mental Health Illness: Eating disorders", 
"Mental Health Illness: Impulse control and addiction disorders", "Mental Health Illness: Personality disorders", "Mental Health Illness: Obsessive-compulsive disorder (OCD)", "Mental Health Illness: Post-traumatic stress disorder (PTSD)", "Mental Health Illness: Schizophrenia",

"Emotion Score: Joy", "Emotion Score: Sadness", "Emotion Score: Anger", "Emotion Score: Fear", "Emotion Score: Surprise",

"Support: Family", "Support: Friends", "Support: Online Communities", "Support: Community Support Services", "Support: Religious or Spiritual Community", "Support: Pets/Emotional Support Animals",

"Stigma: Self", "Stigma: Social", "Stigma: Family", "Stigma: Institutional", "Stigma: Religious/Cultural",

"Coping: Lifestyle Changes", "Coping: Mindfulness and Meditation", "Coping: Hobbies", "Coping: Problem-Solving", "Coping: Emotional Suppression", "Coping: Aggression", "Coping: Social Withdrawal",

"Treatment: Psychotherapy", "Treatment: Medication", "Treatment: Hospitalization", "Treatment: Brain Stimulation Therapies", "Treatment: Complementary & Alternative Medicine", "Treatment: Trauma-Focused Therapies",

"Cause: Trauma and Abuse", "Cause: Grief and Loss", "Cause: Social and Interpersonal Factors", "Cause: Academic and Occupational Stress", "Cause: Financial Strain", "Cause: Legal Issues", "Cause: Biological and Health Factors", "Cause: Substance Use",

"Recovery Status",

"PHQ9: Little Interest", "PHQ9: Feeling Down", "PHQ9: Sleep Problems", "PHQ9: Low Energy", "PHQ9: Appetite Issues", "PHQ9: Low Self-Worth", "PHQ9: Trouble Concentrating", "PHQ9: Restlessness", "PHQ9: Suicidal Thoughts",

"GAD7: Feeling Nervous", "GAD7: Can’t Control Worrying", "GAD7: Worrying Too Much", "GAD7: Trouble Relaxing", "GAD7: Restlessness", "GAD7: Easily Annoyed", "GAD7: Fear Something Awful",
"Hook Present"]
Respond with a valid JSON object containing exactly those 61 fields.
If a value is not known or unclear from the transcript, set it to "Not mentioned".

1. What is the Gender (Male, Female, Unknown) of the person in the video:
2. What is the Age Group (Teen, Adult, Middle-aged, Senior, Unknown) of that person:
3. Based on the transcript (and if applicable, any descriptions provided), infer the person's perceived race. Choose only one from the following options:
African American, White, Black, Asian, Other, Unknown.
4. Based on the Transcript what are the mental Illness the speaker experienced? 
for each type of support listed below, indicate whether it was “Yes” (mentioned or clearly implied) or “No” (not mentioned or unclear)

Anxiety Disorders: Intense fear or worry triggered by certain situations, often with physical symptoms like sweating or rapid heartbeat, that disrupt daily life.

Mood Disorders: Conditions that cause ongoing sadness, extreme happiness, or mood swings, including depression and bipolar disorder.

Psychotic Disorders: Involve losing touch with reality, experiencing hallucinations or false beliefs, as seen in schizophrenia.

Eating Disorders: Unhealthy relationships with food and body image, such as anorexia, bulimia, and binge eating.

Impulse Control & Addiction Disorders: Inability to resist harmful urges like stealing, gambling, or substance use, often leading to serious consequences.

Personality Disorders: Rigid, unhealthy thinking and behavior patterns that interfere with daily life and relationships.

Obsessive-Compulsive Disorder (OCD): Repetitive thoughts (obsessions) and behaviors (compulsions) that a person feels driven to perform.

Post-Traumatic Stress Disorder (PTSD): Ongoing emotional distress and flashbacks after experiencing a traumatic event.

Schizophrenia: A severe mental disorder where a person may hear voices, see things that aren’t real, or hold unusual beliefs.
5. Based on the transcript, detect and quantify the intensity of five core emotions using an approach similar to the RoBERTa emotion classification model (e.g., nateraw/bert-base-uncased-emotion).

Instructions:

For each emotion listed below, estimate its presence as a decimal value between 0 and 1.

Base your estimates only on emotional language, tone, and context conveyed in the transcript.

Core Emotions to Evaluate:

Joy

Sadness

Anger

Fear

Surprise

6.Evaluate the transcript and, for each type of support listed below, indicate whether it was “Yes” (mentioned or clearly implied) or “No” (not mentioned or unclear).

Use context clues such as emotional language, gratitude, specific interactions, or institutional mentions to support your assessment.

Social Support Checklist:
a.Family Support
(Emotional or practical help from parents, siblings,Romantic Partner/Spouse, or relatives)

b.Friends

c.Online Communities or Forums
(Reddit groups, Facebook communities, or mental health apps)

d.Community Support Services
(Hotlines, nonprofit aid, shelters, outreach programs)

e.Religious or Spiritual Community
(Faith-based encouragement, pastoral counseling, or group belonging)

f.Pets or Emotional Support Animals
(Companionship and comfort from animals that reduce distress)

7.Based on the transcript of a person sharing their mental health recovery story.
Evaluate whether the person discusses or implies experiences of stigma related to their mental health.

For each of the following five types of stigma, respond only with “Yes” or “No”:

Return “Yes” if the stigma type is explicitly mentioned or can be reasonably inferred based on the person's story.

Return “No” if the stigma type is not mentioned or cannot be inferred.

Definitions of Stigma Types (use these carefully to judge the transcript):

Self-Stigma
The person feels ashamed, embarrassed, inferior, or weak because of their mental illness.
(e.g., internalized negative beliefs, self-blame, feeling "less than.")

Social (Public) Stigma
The person is judged, rejected, insulted, labeled, or excluded by society — including friends, coworkers, neighbors.
(e.g., being called "crazy," "dangerous," or treated unfairly.)

Family Stigma
The person’s family shows denial, shame, rejection, or different treatment due to the mental illness.
(e.g., distancing, criticism, silence, or minimizing the condition.)

Institutional Stigma
The person experiences discrimination, exclusion, or barriers within systems like healthcare, education, employment, or legal systems.
(e.g., being denied treatment, losing jobs, unfair school/workplace policies.)

Religious or Cultural Stigma
The person is judged based on religious or cultural beliefs that frame mental illness as a sin, curse, weakness of character, or lack of faith.
(e.g., being told to "pray harder" instead of seeking medical help.)
8. Your task is to analyze a person's mental health journey — typically from a YouTube video transcript — and identify the coping mechanisms they mention or describe using.

Coping mechanisms are strategies people use to deal with emotional pain, stress, or mental health challenges. Please read through the transcript and, for each item in the list below, indicate “Yes” if the behavior is clearly mentioned or implied, and “No” if it is not.

Use both direct statements and indirect cues to make informed decisions.

Healthy Coping Strategy Categories & Examples:
Lifestyle Changes: Engaging in regular physical activity is highlighted as a proactive coping mechanism that can mitigate stress and improve overall well-being, Creating consistent daily patterns is suggested as a means to provide a sense of control and predictability, thereby reducing stress.

Mindfulness or Meditation: Practices such as mindfulness meditation are recognized for their role in reducing stress and enhancing emotional regulation.

Hobbies: While not explicitly detailed, the article acknowledges the importance of activities that allow for emotional expression, which can include art, music, or writing. Participating in hobbies or leisure activities is recommended to divert attention from stressors and promote positive emotions.

Problem-Solving: Actively addressing and finding solutions to stress-inducing situations is presented as an effective coping strategy.

Unhealthy Coping Strategy Categories & Examples :

Emotional Suppression or Denial: Ignoring or denying emotions can hinder effective stress management and may lead to increased psychological distress.

Aggression or Lashing Out: Expressing stress through anger or violence is highlighted as a destructive coping method that can damage relationships and personal well-being.

Social Withdrawal: Isolating oneself from social interactions.


*OUTPUT INSTRUCTIONS*

Return a single line containing ten comma separated values (no labels, no extra spaces).  
Each value must be *Yes* or *No, corresponding *in order to categories 1 through 10 above.

9. Based on the transcript, identify the type(s) of mental health treatment the person received. This may include the following categories:

a. Psychotherapy: Involves talking with a mental health professional to explore thoughts, feelings, and behaviors. Can be one-on-one, group, or family therapy.
Examples: Cognitive Behavioral Therapy (CBT), Dialectical Behavior Therapy (DBT), Interpersonal Therapy, Counseling.
b. Medication: Involves prescribed psychiatric drugs that help manage symptoms of mental health conditions.
Examples: Antidepressants, Antipsychotics, Mood Stabilizers, Anti-Anxiety Medications, Sleep Aids.
c. Hospitalization: Involves 24/7 medical or psychiatric care in a hospital or facility when a person is at risk or needs stabilization.
Examples: Psychiatric hospitals, intensive outpatient programs, day-treatment centers.
d. Brain-Stimulation Therapies: Use of electrical or magnetic pulses to treat severe depression or other disorders when medications/therapy don’t work.
Examples: Electroconvulsive Therapy (ECT), Transcranial Magnetic Stimulation (TMS), Deep Brain Stimulation (DBS), Vagus Nerve Stimulation (VNS).
e. Complementary & Alternative Medicine (CAM): Non-traditional methods used alongside or instead of standard treatment.
Examples: Acupuncture, herbal supplements, energy healing, massage therapy.
f. Trauma-Focused Therapies: Special therapies to process and heal trauma.
Examples: EMDR (Eye Movement Desensitization and Reprocessing), Trauma-focused CBT, Somatic Experiencing.

Please respond with the specific treatment names or categories mentioned in the transcript, using single words or short phrases.

10. Analyze the transcript to identify and describe the key situations, experiences, and stressors that the individual attributes to the onset or worsening of their mental health condition. These factors may include personal trauma, family dynamics, social or environmental pressures, or biological vulnerabilities. Focus on what the person believes contributed to their mental health struggles, including both specific events and broader life circumstances.

Then, categorize the cause(s) according to the following verified mental health risk domains, which are supported by evidence from global and cross-cultural mental health research:

a. Trauma and Abuse – Exposure to emotionally, physically, or psychologically harmful experiences, including childhood maltreatment, domestic violence, or war-related trauma.
b. Grief and Loss – Emotional distress due to the death of a loved one, significant relationship loss, or disruption of important life attachments.
c. Social and Interpersonal Factors: Mental health challenges often arise from a combination of societal pressures (discrimination, cultural conflict, stigma), interpersonal difficulties (family conflict, peer bullying, social isolation), and major life transitions (relocation, immigration, job loss, or changes in personal identity roles). These factors can individually or collectively contribute to emotional distress and psychological strain.
e. Academic or Occupational Stress – Pressure from school, exams, work responsibilities, or fear of failure.
f. Financial Strain – Ongoing stress or uncertainty due to poverty, debt, job insecurity, or economic inequality.
g. Legal Issues – Mental stress arising from legal problems, detention, or interactions with the justice system.
h.Biological, Genetic, and Health-Related Factors:
Mental health challenges can be influenced by a combination of biological vulnerabilities (including genetic predispositions, neurochemical imbalances, and prenatal complications) and health-related conditions (such as chronic illness, physical disability, or persistent concerns about bodily well-being). These factors may independently or interactively increase the risk of psychological distress.
i. Substance Use – Mental health issues linked to the use or misuse of alcohol, drugs, or prescription medication.

Your response should:
Clearly describe the specific causes as shared in the transcript.

List one or more relevant categories from the list above that best describe those causes.

Avoid Yes/No answers and provide concise labels or phrases (e.g., "childhood trauma," "immigration stress," "job loss").

11. Based on the transcript or narrative where someone shares their mental health experience — such as from a YouTube video or interview — and determine their current mental health status based on what they describe.

Review the content carefully to identify how the person is feeling and functioning at the time of sharing their story. Match what they say to one of the six categories listed below. Use context clues such as statements about their current symptoms, improvements, treatment use, emotions, or mindset.

Select One of the Following as Your Final Answer:
a.Fully Recovered:

The person has overcome their mental health struggles and no longer experiences significant symptoms.

b.Ongoing Recovery:

The person is stable or improving, actively managing their mental health to maintain wellness.

c.Not Recovered:

The person is still facing significant mental health challenges, showing little improvement or experiencing setbacks.

12. Evaluate whether the person describes symptoms that align with the PHQ-9 (depression) and GAD-7 (anxiety) screening tools.
For each symptom listed below, simply answer "Yes" if the symptom is clearly mentioned or implied in their story, or "No" if it is not mentioned or unclear.

Rules:

"Mentioned or clearly implied" → Yes

"Not mentioned or unclear" → No

PHQ-9 Symptoms (Depression):

PHQ9: Little Interest (loss of interest in hobbies, socializing, work)

PHQ9: Feeling Down (sadness, hopelessness, low mood)

PHQ9: Sleep Problems (insomnia, oversleeping, trouble staying asleep)

PHQ9: Low Energy (fatigue, feeling drained, no motivation)

PHQ9: Appetite Issues (loss of appetite, overeating, weight changes)

PHQ9: Low Self-Worth (self-criticism, guilt, feeling like a burden)

PHQ9: Trouble Concentrating (difficulty focusing or making decisions)

PHQ9: Restlessness (physical agitation, feeling tense or jittery)

PHQ9: Suicidal Thoughts (thoughts of death, self-harm, suicide ideation)

GAD-7 Symptoms (Anxiety):

GAD7: Feeling Nervous (frequent worry, feeling on edge)

GAD7: Can’t Control Worrying (worry spirals, intrusive anxious thoughts)

GAD7: Worrying Too Much (about everyday things, worst-case scenarios)

GAD7: Trouble Relaxing (inability to unwind, constant tension)

GAD7: Restlessness (fidgeting, pacing, can't sit still)

GAD7: Easily Annoyed (irritability, short temper, emotional outbursts)

GAD7: Fear Something Awful (anticipation of bad events, paranoia, panic)

13. Carefully examine the first 5-30 seconds of the transcript and determine whether the video contains a hook — an attention-grabbing statement or moment meant to make the viewer want to keep watching.

Hook Text: Quote or describe what happens in the first 15 seconds.

Describe the Hook: In 1–2 sentences, explain what makes this a hook (e.g., Does it start with a personal story? Use a strong emotion? Ask a surprising question? Present a bold statement? Offer a solution or promise a benefit? Show quick visuals or give a preview? Something else?)  You may refer to examples like:

Examples of possible hooks include:

Starting with a personal story preview

Sharing an emotional trigger

Asking a question or sparking curiosity

Making a bold or surprising statement

Presenting a problem and hinting at a solution

Promising value or a clear benefit to the viewer

Showing a quick visual tease or a fast-paced summary

If yes, provide the following:

Hook Present (Yes/No):


Transcript:
{transcript}
"""
    generation_config = GenerationConfig(
    temperature=0.0)
    model = genai.GenerativeModel("gemini-2.0-flash",generation_config=generation_config)
    for attempt in range(3):
        try:
            response = model.generate_content(prompt)
            if not response.text.strip():  # ✅ check for empty
                print("❌ Empty response from Gemini. Skipping.")
                return None
            return response.text.strip()

        except Exception as e:
            if "429" in str(e):
                print("⏳ Rate limit hit. Waiting 60 seconds...")
                time.sleep(60)
            else:
                print(f"❌ Gemini error: {e}")
                return None
    return None

def parse_gemini_response(analysis: str) -> list:
    try:
        # Extract just the JSON from a response block (in case it's wrapped in triple backticks)
        if analysis.strip().startswith("json"):
            analysis = analysis.strip().strip("json").strip("").strip()

        parsed_json = json.loads(analysis)
        return [parsed_json.get(key, "Not mentioned") for key in append_to_excel.__annotations__["data_row"]]
    except Exception as e:
        print(f"❌ Error parsing JSON: {e}")
        return ["Not mentioned"] * 124

def summarize_youtube_videos(video_urls):
    for url in video_urls:
        try:
            print(f"\n🔍 Analyzing: {url}")
            video_id = url.split("v=")[-1].split("&")[0].split("?si=")[0].split("/")[-1]

            try:
                transcript = get_youtube_transcript(video_id)
            except (TranscriptsDisabled, NoTranscriptFound):
                print("❗ Transcript not available. Attempting fallback...")
                # transcript = download_and_transcribe(url)

            if not transcript:
                print(f"❌ No transcript for: {url}")
                continue

            # Get emotion scores before calling Gemini (if still needed)
            emotion_label, emotion_scores = get_emotion_scores(transcript)

            # yt = YouTube(url)
            # video_title = yt.title.strip()


            # Get structured JSON response from Gemini
            analysis = analyze_transcript(transcript, url)
            if not analysis:
                print(f"❌ No analysis for {url}")
                continue

            print(f"\n📄 Extracted Info:\n{analysis}")

            try:
                # Clean up triple backtick-wrapped JSON if needed
                cleaned_json = analysis.strip()
                if cleaned_json.startswith("```json"):
                    cleaned_json = cleaned_json.removeprefix("```json").removesuffix("```").strip()
                elif cleaned_json.startswith("```"):
                    cleaned_json = cleaned_json.removeprefix("```").removesuffix("```").strip()


                parsed_json = json.loads(cleaned_json)

                # Ensure exact 124 columns, fill in 'Not mentioned' for any missing keys
                # data_row = [parsed_json.get(col, "Not mentioned") for col in COLUMNS_124]
                data_row = [url] + [parsed_json.get(col, "Not mentioned") for col in COLUMNS_124 if col != "YouTube URL"]


                print(f"📊 Parsed ({len(data_row)} fields): {data_row}")
                append_to_excel(EXCEL_OUTPUT_FILE, data_row)

            except Exception as json_err:
                print(f"❌ JSON parsing error: {json_err}")
                continue

        except Exception as e:
            print(f"⚠ Error processing {url}: {e}")
        finally:
            print("⏱ Waiting 5 seconds before next video...")
            time.sleep(5)


# def download_and_transcribe(url):
#     try:
#         yt = YouTube(url)
#         video_title = yt.title.replace(" ", "_")
#         video_path = yt.streams.filter(only_audio=True).first().download(filename=f"{video_title}.mp4")
#         audio_path = f"{video_title}.mp3"
#         clip = VideoFileClip(video_path)
#         clip.audio.write_audiofile(audio_path)
#         clip.close()
#         os.remove(video_path)
#         os.remove(audio_path)
#         return None
#     except Exception as e:
#         print(f"Error downloading/transcribing {url}: {e}")
#         return None

if __name__ == "__main__":

    video_urls = load_video_urls_from_excel(EXCEL_INPUT_FILE, EXCEL_INPUT_SHEET)
    summarize_youtube_videos(video_urls)